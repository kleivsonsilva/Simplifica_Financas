"""
REPORT SERVICE - Microsserviço de Relatórios e Exportação
Responsabilidade: Geração de relatórios, exportação Excel/PDF, análises
Banco: MongoDB (leitura de transações)
Cache: Redis (relatórios gerados em cache)
Porta: 5004
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from datetime import datetime
from functools import wraps
from pymongo import MongoClient
import redis
import jwt
import json
import os
import io
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s [REPORT] %(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

MONGO_URL  = os.getenv('MONGO_URL')
REDIS_URL  = os.getenv('REDIS_URL')
JWT_SECRET = os.getenv('JWT_SECRET')
PORT       = int(os.getenv('SERVICE_PORT', 5004))

# ─── VALIDAÇÃO DE VARIÁVEIS CRÍTICAS ─────────────────────────
_missing = [v for v, val in [('MONGO_URL', MONGO_URL), ('JWT_SECRET', JWT_SECRET)] if not val]
if _missing:
    raise EnvironmentError(f"Variáveis de ambiente obrigatórias não definidas: {_missing}")

def get_mongo():
    client = MongoClient(MONGO_URL, serverSelectionTimeoutMS=10000)
    return client['simplifica_financas']

def get_redis():
    if not REDIS_URL:
        return None
    return redis.from_url(REDIS_URL, decode_responses=True)

# ─── AUTENTICAÇÃO COM BLACKLIST ────────────────────────────────
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token:
            return jsonify({'error': 'Token não fornecido'}), 401
        try:
            r = get_redis()
            if r and r.get(f"blacklist:{token}"):
                return jsonify({'error': 'Token revogado'}), 401
            request.user = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
        except jwt.ExpiredSignatureError:
            return jsonify({'error': 'Token expirado'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'error': 'Token inválido'}), 401
        return f(*args, **kwargs)
    return decorated

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'service': 'report-service'})

@app.route('/relatorios/resumo', methods=['GET'])
@token_required
def resumo():
    user_id   = request.user['sub']
    mes       = request.args.get('mes')
    cache_key = f"relatorio:{user_id}:{mes or 'all'}"

    try:
        r      = get_redis()
        if r:
            cached = r.get(cache_key)
            if cached:
                return jsonify(json.loads(cached))

        db     = get_mongo()
        filtro = {'usuario_id': user_id, 'ativo': True}
        if mes:
            filtro['data'] = {'$regex': f'^{mes}'}

        geral = list(db.transacoes.aggregate([
            {'$match': filtro},
            {'$group': {'_id': '$tipo', 'total': {'$sum': '$valor'}, 'count': {'$sum': 1}}}
        ]))

        por_categoria = list(db.transacoes.aggregate([
            {'$match': {**filtro, 'tipo': 'despesa'}},
            {'$group': {'_id': '$categoria', 'total': {'$sum': '$valor'}}},
            {'$sort': {'total': -1}},
            {'$limit': 10}
        ]))

        receitas = next((g['total'] for g in geral if g['_id'] == 'receita'), 0)
        despesas = next((g['total'] for g in geral if g['_id'] == 'despesa'), 0)

        resultado = {
            'receitas':         float(receitas),
            'despesas':         float(despesas),
            'saldo':            float(receitas - despesas),
            'maior_categoria':  por_categoria[0]['_id'] if por_categoria else None,
            'por_categoria':    [{'categoria': c['_id'], 'total': float(c['total'])} for c in por_categoria],
            'gerado_em':        datetime.utcnow().isoformat()
        }

        if r:
            r.setex(cache_key, 600, json.dumps(resultado))
        return jsonify(resultado)

    except Exception as e:
        logger.error(f"❌ Erro no resumo: {e}")
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/relatorios/exportar/excel', methods=['GET'])
@token_required
def exportar_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl não instalado'}), 500

    user_id = request.user['sub']
    mes     = request.args.get('mes')

    try:
        db         = get_mongo()
        filtro     = {'usuario_id': user_id, 'ativo': True}
        if mes:
            filtro['data'] = {'$regex': f'^{mes}'}
        transacoes = list(db.transacoes.find(filtro).sort('data', -1))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Transações'

        cabecalho   = ['Data', 'Tipo', 'Descrição', 'Categoria', 'Valor (R$)']
        header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        for col, titulo in enumerate(cabecalho, 1):
            cell           = ws.cell(row=1, column=col, value=titulo)
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = 20

        for row, t in enumerate(transacoes, 2):
            ws.cell(row=row, column=1, value=t.get('data', ''))
            ws.cell(row=row, column=2, value=t.get('tipo', '').title())
            ws.cell(row=row, column=3, value=t.get('descricao', ''))
            ws.cell(row=row, column=4, value=t.get('categoria', ''))
            cell              = ws.cell(row=row, column=5, value=float(t.get('valor', 0)))
            cell.number_format = 'R$ #,##0.00'
            cell.font          = Font(color='DC2626' if t.get('tipo') == 'despesa' else '16A34A')

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'simplifica_financas_{mes or "completo"}.xlsx'
        )

    except Exception as e:
        logger.error(f"❌ Erro ao exportar Excel: {e}")
        return jsonify({'error': 'Erro ao gerar Excel'}), 500

@app.route('/relatorios/exportar/pdf', methods=['GET'])
@token_required
def exportar_pdf():
    try:
        from fpdf import FPDF
    except ImportError:
        return jsonify({'error': 'fpdf não instalado'}), 500

    user_id = request.user['sub']
    mes     = request.args.get('mes')

    try:
        db         = get_mongo()
        filtro     = {'usuario_id': user_id, 'ativo': True}
        if mes:
            filtro['data'] = {'$regex': f'^{mes}'}
        transacoes = list(db.transacoes.find(filtro).sort('data', -1).limit(100))
        receitas   = sum(t['valor'] for t in transacoes if t['tipo'] == 'receita')
        despesas   = sum(t['valor'] for t in transacoes if t['tipo'] == 'despesa')

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, 'Simplifica Financas - Relatorio Financeiro', ln=True, align='C')
        pdf.set_font('Arial', '', 10)
        pdf.cell(0, 6, f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', ln=True, align='C')
        pdf.ln(5)

        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 8, 'RESUMO', ln=True)
        pdf.set_font('Arial', '', 11)
        pdf.cell(0, 7, f'Total de Receitas: R$ {receitas:,.2f}', ln=True)
        pdf.cell(0, 7, f'Total de Despesas: R$ {despesas:,.2f}', ln=True)
        pdf.cell(0, 7, f'Saldo: R$ {(receitas - despesas):,.2f}', ln=True)
        pdf.ln(5)

        pdf.set_font('Arial', 'B', 10)
        pdf.set_fill_color(99, 102, 241)
        pdf.set_text_color(255, 255, 255)
        for col, w in [('Data', 30), ('Tipo', 25), ('Descricao', 75), ('Categoria', 35), ('Valor', 25)]:
            pdf.cell(w, 8, col, fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Arial', '', 9)

        for i, t in enumerate(transacoes):
            pdf.set_fill_color(240, 240, 255) if i % 2 == 0 else pdf.set_fill_color(255, 255, 255)
            pdf.cell(30, 7, str(t.get('data', '')),             fill=True)
            pdf.cell(25, 7, t.get('tipo', '').title(),          fill=True)
            pdf.cell(75, 7, str(t.get('descricao', ''))[:35],   fill=True)
            pdf.cell(35, 7, t.get('categoria', ''),             fill=True)
            pdf.cell(25, 7, f'R${t.get("valor", 0):,.2f}',      fill=True)
            pdf.ln()

        pdf_bytes = pdf.output(dest='S').encode('latin-1')
        output    = io.BytesIO(pdf_bytes)

        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'simplifica_financas_{mes or "completo"}.pdf'
        )

    except Exception as e:
        logger.error(f"❌ Erro ao exportar PDF: {e}")
        return jsonify({'error': 'Erro ao gerar PDF'}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=PORT, debug=False)
